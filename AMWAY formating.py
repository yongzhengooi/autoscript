import subprocess
from urllib.parse import urlparse
import re
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
import platform
#variables
file_name=None
current_row=1

def print_lines():
	print("==========================================")

def create_row(workbook,data,row,increase=False,numbering=True):
	global current_row
	for i, value in enumerate(data, start=1):
		workbook.cell(row, column=i, value=value)
	if numbering:
		workbook.cell(row,1,current_row-1)

	if increase:
		current_row+=1

def mod_scope(workbook,row,column,value):
	workbook.cell(row,column,value)

def format(sheet):
	font = Font(name="Calibri", size=10)
	alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
	border = Border(
		top=Side(border_style="thin", color="808080"),
		left=Side(border_style="thin", color="808080"),
		right=Side(border_style="thin", color="808080"),
		bottom=Side(border_style="thin", color="808080")
	)

	for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
		for cell in row:
			cell.font=font
			cell.alignment=alignment
			cell.border=border

def format_header(sheet):
	font = Font(name="Calibri", size=10, bold=True)
	for cell in sheet:
		cell.font=font


def format_severity(sheet):
	red_fill = PatternFill(fill_type="solid", fgColor="D99594")     # Red
	orange_fill = PatternFill(fill_type="solid", fgColor="FABF8F")  # Orange
	yellow_fill = PatternFill(fill_type="solid", fgColor="FAF6B0")  # Yellow
	blue_fill = PatternFill(fill_type="solid", fgColor="DBE5F1")
	green_fill = PatternFill(fill_type="solid", fgColor="C2D69B")

	header_row = [cell.value for cell in sheet[1]]
	try:
    		severity_col_index = header_row.index("Severity") + 1
    		cvss_col_index = header_row.index("Comments") + 1
	except ValueError:
    		print ("Column 'Severity' and 'Comments' not found.")

	for row in range(2, sheet.max_row + 1):
		severity_cell = sheet.cell(row=row, column=severity_col_index)
		cvss_cell = sheet.cell(row=row, column=cvss_col_index)
		value = str(cvss_cell.value).strip().upper().split("\n")
		cvss_value=str(value[2]).split(": ")
		cvss_score=float(cvss_value[1])
		if cvss_score >=9 and cvss_score <=10:
			severity_cell.fill = red_fill
			severity_cell.value = "Critical"
		elif cvss_score >= 7 and cvss_score <= 8.9:
			severity_cell.fill = orange_fill
			severity_cell.value = "High"
		elif cvss_score >=4 and cvss_score <= 6.9:
			severity_cell.fill = yellow_fill
			severity_cell.value = "Medium"
		elif cvss_score >=1 and cvss_score <= 3.9:
			severity_cell.fill = green_fill
			severity_cell.value = "Low"
		else:
			severity_cell.fill = blue_fill
			severity_cell.value = "Info"

def add_port_column(sheet):
	headers = [cell.value for cell in sheet[1]]
	if "Port" not in headers:
		#country=filename.split("_")
		sheet.insert_cols(scope_index )
		sheet.cell(row=1, column=scope_index ).value = "Port"
		for row in range(2, sheet.max_row ):
    			sheet.cell(row=row, column=scope_index).value = ""

def column_width(sheet):
	sheet.column_dimensions['A'].width = 5
	sheet.column_dimensions['B'].width = 15
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 40
	sheet.column_dimensions['E'].width = 40
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 40
	sheet.column_dimensions['H'].width = 15
	sheet.column_dimensions['I'].width = 60
#Find xlsx
#print_lines()
#print('Searching for xlsx files in current folder:')
#command='find . -type f -name "*xlsx" 2>/dev/null'
#result = subproce
system = platform.system()

if system == "Windows":
	print('Seaching for xlsx files in current folder ')
	result = subprocess.run(['cmd', '/c', 'dir', '/b', '*.xlsx'], capture_output=True, text=True)
	print(result.stdout)
	file_name=str(result.stdout).strip().split("\n")
	#print(file_name)
elif system == "Linux":
	print('Searching for xlsx files in current folder:')
	command='find . -type f -name "*xlsx" -maxdepth 1 2>/dev/null'
	result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, shell=True)
	print(result.stdout)
	file_name=str(result.stdout).strip().split("\n")

#Search for scope cordinate
def main(workbook_location):
	scope_location = None
	last_row = None
	verfied=False
	global scope_index
	global current_row
	wb = load_workbook(workbook_location)
	sheet = wb.active
	print_lines()
	print('Location index for "Scope" ')
	for cell in sheet[1]:
		if "scope" in str(cell.value).lower():
			scope_location = cell.coordinate[0]
			scope_index=(ord(scope_location)-64)
			break

	if scope_location:
		print(f"Column locate at: {scope_location}")
		last_row = sheet.max_row
		while all(cell.value is None for cell in sheet[last_row]):
			last_row -= 1
		print(f"Last rows number: {last_row}")
		verfied=True
		add_port_column(sheet)

	#List scope values
	if verfied:
		#create new excel
		workbook=openpyxl.Workbook()
		newexcel=workbook.active
		header=[cell.value for cell in sheet[1]] #header
		print(header)
		create_row(newexcel,header,current_row,increase=True,numbering=False) #header
		print_lines()
		print("Stripping the scope to excel ....")
		for i in range(2,last_row +1):
			row_values =[cell.value for cell in sheet[i]]
			scope=row_values[-1].strip()
			print(scope)
			scope_list=scope.split('\n')
			if len(scope_list) >=1:
				for port in scope_list:
					if "TCP" in port or "UDP" in port:
						#print("TCP found")
						#print(port)
						port_num = port.strip().split(': ')
						#print(port_num[0])
						print('\n')
						port_scope=port_num[1].split(', ')
						#print(port_scope)
						#print(scope-index)
						for list_index in range (0,len(port_scope)):
							create_row(newexcel,row_values,current_row)
							mod_scope(newexcel,current_row,(scope_index)+1,port_scope[list_index])
							mod_scope(newexcel,current_row,8,str(port_num[0]).replace("/",":"))
							current_row+=1
					else:
						create_row(newexcel,row_values,current_row)
						mod_scope(newexcel,current_row,scope_index,port)
						if "http" in port:
							if re.search(r"[:,]\s*(\d+)", port):
								match=re.findall(r"[:,]\s*(\d+)", port)
								if match:
									url=port.rsplit(":",1)[0]
									url_path=urlparse(port)
									path=str(url_path.path).replace("_x000D_", "")
									for h_port in match:
										new_url=url+":"+h_port+path
										create_row(newexcel,row_values,current_row)
										mod_scope(newexcel,current_row,scope_index+1,new_url)
										mod_scope(newexcel,current_row,8,"TCP:"+h_port)
										current_row+=1
							elif "https:" in port:
								mod_scope(newexcel,current_row,8,"TCP:443")
								current_row+=1
							elif "http:" in port:
								mod_scope(newexcel,current_row,8,"TCP:80")
								current_row+=1
			else:
				create_row(newexcel,row_values,current_row,increase=True)
		
		print("Complete stripping ...")
		print_lines()
		format_header(newexcel[1])
		print_lines()
		print("Coloring the severity ...")
		format_severity(newexcel)
		print_lines()
		print("Adjusting column width ....")
		column_width(newexcel)
		print("completed")
		print("Configure the excel format ...")
		format(newexcel)
		if system == "Linux":
			file_split=str(workbook_location).split("/")
			file_path = './new/'+file_split[1]
		else:
			file_path = './new/'+workbook_location
		#Delete row num and affected host
		newexcel.delete_cols(1)
		#newexcel.delete_cols(9)
		workbook.save(file_path)


#MAIN
if not os.path.exists("new"):
    os.makedirs("new")

try:
	for file in file_name:
		print("\n\n")
		print(f"Creating excel for {file}")
		current_row=1
		scope_index=None
		main(file)
except Exception as e:
	print(e)

print("Amway Mr. Chan Excel format semi automator - don't forget to fix the formating, merge common cell and for outdated component need to specify affected component manually")
